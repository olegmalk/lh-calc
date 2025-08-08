#!/usr/bin/env python3
"""
Extract all formulas, relations, and data from Excel calculator
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import json
import re
from collections import defaultdict
from pathlib import Path

def extract_excel_data(filepath):
    """Extract comprehensive data from Excel file"""
    
    print(f"Loading Excel file: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=False, keep_vba=True)
    
    # Also load with data_only=True to get calculated values
    wb_values = openpyxl.load_workbook(filepath, data_only=True)
    
    result = {
        'sheets': {},
        'formulas': {},
        'named_ranges': {},
        'relations': defaultdict(list),
        'data_validations': {},
        'summary': {}
    }
    
    # Extract named ranges
    print("\n=== NAMED RANGES ===")
    for name, coord in wb.defined_names.items():
        result['named_ranges'][name] = str(coord)
        print(f"{name}: {coord}")
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        print(f"\n=== SHEET: {sheet_name} ===")
        ws = wb[sheet_name]
        ws_values = wb_values[sheet_name]
        
        sheet_data = {
            'cells': {},
            'formulas': {},
            'values': {},
            'merged_cells': [],
            'data_validations': []
        }
        
        # Get merged cells
        for merged_range in ws.merged_cells.ranges:
            sheet_data['merged_cells'].append(str(merged_range))
        
        # Get data validations
        for dv in ws.data_validations.dataValidation:
            if dv.cells:
                sheet_data['data_validations'].append({
                    'type': dv.type,
                    'formula1': dv.formula1,
                    'formula2': dv.formula2,
                    'cells': str(dv.cells)
                })
        
        # Process all cells
        formula_count = 0
        data_count = 0
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                    
                    # Store formula if exists
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formula = cell.value
                        sheet_data['formulas'][cell_ref] = formula
                        formula_count += 1
                        
                        # Extract cell references from formula
                        # Pattern to match cell references like A1, $A$1, Sheet1!A1
                        pattern = r'(?:[\w]+!)?(?:\$)?[A-Z]+(?:\$)?[0-9]+'
                        references = re.findall(pattern, formula)
                        
                        if references:
                            result['relations'][f"{sheet_name}!{cell_ref}"] = references
                        
                        # Get calculated value
                        calc_value = ws_values[cell_ref].value
                        sheet_data['values'][cell_ref] = {
                            'formula': formula,
                            'calculated': calc_value
                        }
                        
                        print(f"  Formula at {cell_ref}: {formula[:50]}... = {calc_value}")
                    else:
                        # Store regular data
                        sheet_data['cells'][cell_ref] = cell.value
                        data_count += 1
                        
                        # Check if it's a header or important label
                        if isinstance(cell.value, str) and len(cell.value) > 2:
                            if cell.font and cell.font.bold:
                                print(f"  Header at {cell_ref}: {cell.value}")
        
        result['sheets'][sheet_name] = sheet_data
        print(f"  Found {formula_count} formulas and {data_count} data cells")
    
    # Create summary
    total_formulas = sum(len(sheet['formulas']) for sheet in result['sheets'].values())
    total_sheets = len(result['sheets'])
    
    result['summary'] = {
        'total_sheets': total_sheets,
        'total_formulas': total_formulas,
        'sheet_names': list(wb.sheetnames),
        'has_vba': wb.vba_archive is not None if hasattr(wb, 'vba_archive') else False
    }
    
    print(f"\n=== SUMMARY ===")
    print(f"Total sheets: {total_sheets}")
    print(f"Total formulas: {total_formulas}")
    print(f"Sheet names: {', '.join(wb.sheetnames)}")
    
    return result

def analyze_formulas(data):
    """Analyze formulas to understand calculation logic"""
    
    print("\n=== FORMULA ANALYSIS ===")
    
    formula_types = defaultdict(list)
    functions_used = defaultdict(int)
    
    for sheet_name, sheet_data in data['sheets'].items():
        for cell_ref, formula in sheet_data['formulas'].items():
            # Extract functions used
            func_pattern = r'([A-Z]+)\('
            functions = re.findall(func_pattern, formula)
            for func in functions:
                functions_used[func] += 1
            
            # Categorize formula
            if 'IF' in formula:
                formula_types['conditional'].append(f"{sheet_name}!{cell_ref}")
            elif 'SUM' in formula or 'СУММ' in formula:
                formula_types['summation'].append(f"{sheet_name}!{cell_ref}")
            elif 'VLOOKUP' in formula or 'ВПР' in formula:
                formula_types['lookup'].append(f"{sheet_name}!{cell_ref}")
            elif '*' in formula or '/' in formula:
                formula_types['calculation'].append(f"{sheet_name}!{cell_ref}")
            else:
                formula_types['other'].append(f"{sheet_name}!{cell_ref}")
    
    print("\nFormula Categories:")
    for category, cells in formula_types.items():
        print(f"  {category}: {len(cells)} formulas")
    
    print("\nTop Functions Used:")
    for func, count in sorted(functions_used.items(), key=lambda x: x[1], reverse=True)[:10]:
        print(f"  {func}: {count} times")
    
    return formula_types, functions_used

def save_results(data, output_path):
    """Save extracted data to JSON file"""
    
    # Convert defaultdict to regular dict for JSON serialization
    data['relations'] = dict(data['relations'])
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    
    print(f"\nResults saved to: {output_path}")

def create_documentation(data, output_path):
    """Create markdown documentation of the Excel structure"""
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("# Excel Calculator Structure Documentation\n\n")
        f.write("## Summary\n")
        f.write(f"- Total Sheets: {data['summary']['total_sheets']}\n")
        f.write(f"- Total Formulas: {data['summary']['total_formulas']}\n")
        f.write(f"- Sheets: {', '.join(data['summary']['sheet_names'])}\n\n")
        
        f.write("## Sheets Analysis\n\n")
        
        for sheet_name, sheet_data in data['sheets'].items():
            f.write(f"### {sheet_name}\n")
            f.write(f"- Formulas: {len(sheet_data['formulas'])}\n")
            f.write(f"- Data Cells: {len(sheet_data['cells'])}\n")
            
            if sheet_data['merged_cells']:
                f.write(f"- Merged Cells: {len(sheet_data['merged_cells'])}\n")
            
            # Document key formulas
            if sheet_data['formulas']:
                f.write("\n#### Key Formulas:\n")
                for i, (cell, formula) in enumerate(list(sheet_data['formulas'].items())[:10]):
                    f.write(f"- `{cell}`: `{formula[:100]}{'...' if len(formula) > 100 else ''}`\n")
                
                if len(sheet_data['formulas']) > 10:
                    f.write(f"- ... and {len(sheet_data['formulas']) - 10} more formulas\n")
            
            f.write("\n")
        
        # Document relationships
        if data['relations']:
            f.write("## Cell Dependencies\n\n")
            for source, targets in list(data['relations'].items())[:20]:
                f.write(f"- `{source}` depends on: {', '.join(targets[:5])}{'...' if len(targets) > 5 else ''}\n")
            
            if len(data['relations']) > 20:
                f.write(f"- ... and {len(data['relations']) - 20} more dependencies\n")
    
    print(f"Documentation saved to: {output_path}")

if __name__ == "__main__":
    excel_path = "/home/vmuser/dev/lh_calc/Себестоимость ТЕПЛОБЛОК шаблон для снабжения версия 7 ДЛЯ БИТРИКС (1) (1).xlsx"
    
    # Extract data
    data = extract_excel_data(excel_path)
    
    # Analyze formulas
    formula_types, functions_used = analyze_formulas(data)
    
    # Save results
    output_dir = Path("/home/vmuser/dev/lh_calc/lh-calculator")
    save_results(data, output_dir / "excel_formulas.json")
    create_documentation(data, output_dir / "EXCEL_STRUCTURE.md")
    
    print("\n=== EXTRACTION COMPLETE ===")
    print("Check excel_formulas.json for detailed data")
    print("Check EXCEL_STRUCTURE.md for documentation")